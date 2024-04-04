import openpyxl as px
import os
from django.utils import timezone
from django.db.models import Sum

from db.models import Organization, PaymentType, FuelType, FuelStorage, FuelColumnPointer, SaleFuel, \
    OrganizationFuelColumns, OrganizationFuelTypes, FuelPrice, FuelStorageHistory


def get_report_xlsx(user, week=False, fuel_type=None):
    file_path = os.path.join(os.getcwd(), 'static', 'report.xlsx')
    wb = px.load_workbook(file_path)
    ws = wb.active
    ws.title = "Ҳафталик ҳисобот" if week else "Ойлик ҳисобот"
    organ_fuel_columns = OrganizationFuelColumns.objects.filter(organization=user.organization).count()
    residual = FuelStorage.objects.filter(organization=user.organization, fuel_type=fuel_type, is_over=False,
                                          created_at__lt=timezone.now()).aggregate(total=Sum('residual'))['total']
    i = 6
    date_range_all = [timezone.now().date() - timezone.timedelta(days=day) for day in range(7 if week else 30)]
    storage_history = FuelStorageHistory.objects.filter(organization=user.organization, fuel_type=fuel_type,
                                                        created_at__date__in=date_range_all)
    begin, end = storage_history.first().begin, storage_history.first().end
    for date in date_range_all[::-1]:
        ws.merge_cells(f'B{i}:B{i + organ_fuel_columns - 1}')
        ws[f'B{i}'] = f"{date.day}-{date.strftime('%B')[0:5]}"
        ws.merge_cells(f'C{i}:C{i + organ_fuel_columns - 1}')
        ws[f'C{i}'] = begin
        ws.merge_cells(f'D{i}:D{i + organ_fuel_columns - 1}')
        date_fuel_storage = FuelStorage.objects.filter(organization=user.organization, fuel_type=fuel_type,
                                                       created_at__date=date).last()
        ws[f'D{i}'] = date_fuel_storage.remainder if date_fuel_storage else None
        pointer = FuelColumnPointer.objects.filter(organ=user.organization,
                                                   created_at__date=date)
        item = 0
        for point in pointer:
            ws[f'E{i + item}'] = point.size_last
            ws[f'F{i + item}'] = point.size_first
            item += 1
        sale_fuel = SaleFuel.objects.filter(organization=user.organization, fuel_type=fuel_type, created_at__date=date)
        fuel_price_last = FuelPrice.objects.filter(organization=user.organization, fuel_type=fuel_type,
                                                   created_at__lt=date).last()
        ws.merge_cells(f'G{i}:G{i + organ_fuel_columns - 1}')
        size = sale_fuel.aggregate(total=Sum('cash_size'))['total'] + sale_fuel.aggregate(
            total=Sum('card_size'))['total'] if sale_fuel.exists() else None
        ws[f'G{i}'] = size if sale_fuel.exists() else None
        ws.merge_cells(f'H{i}:H{i + organ_fuel_columns - 1}')
        ws[f'H{i}'] = (sale_fuel.first().price - sale_fuel.first().benefit) / size if sale_fuel.exists() else None
        ws.merge_cells(f'I{i}:I{i + organ_fuel_columns - 1}')
        ws[f'I{i}'] = fuel_price_last.price if fuel_price_last else None
        ws.merge_cells(f'J{i}:J{i + organ_fuel_columns - 1}')
        ws[f'J{i}'] = sale_fuel.first().price if sale_fuel.exists() else None
        ws.merge_cells(f'K{i}:K{i + organ_fuel_columns - 1}')
        ws[f'K{i}'] = sale_fuel.first().card_size if sale_fuel.exists() else None
        ws.merge_cells(f'L{i}:L{i + organ_fuel_columns - 1}')
        ws[f'L{i}'] = sale_fuel.first().card_size * fuel_price_last.price if sale_fuel.exists() else None
        ws.merge_cells(f'M{i}:M{i + organ_fuel_columns - 1}')
        ws[f'M{i}'] = sale_fuel.first().cash_size if sale_fuel.exists() else None
        ws.merge_cells(f'N{i}:N{i + organ_fuel_columns - 1}')
        ws[f'N{i}'] = sale_fuel.first().cash_size * fuel_price_last.price if sale_fuel.exists() else None
        ws.merge_cells(f'O{i}:O{i + organ_fuel_columns - 1}')
        ws[f'O{i}'] = sale_fuel.first().benefit if sale_fuel.exists() else None
        ws.merge_cells(f'P{i}:P{i + organ_fuel_columns - 1}')
        ws[f'P{i}'] = end
        history = FuelStorageHistory.objects.filter(organization=user.organization, fuel_type=fuel_type,
                                                    created_at__date=date).first()
        if history:
            begin, end = history.begin, history.end
        i += organ_fuel_columns

    path = os.path.join(os.getcwd(), 'static', 'crm_report.xlsx')
    wb.save(path)
    return path
